"""Generate TAMS-branded XLSX reports with financial data tables and charts."""

import os
import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from templates.report_structure import SECTION_ORDER, SECTION_TITLES, report_filename


# --- Brand colors (TAM Liquid Glass palette) ---
DARK_BLUE = "222F62"        # TAM Deep Blue (primary brand)
ACCENT_BLUE = "1A6DB6"      # TAM Light Blue (accent)
TURQUOISE = "6CB9B6"        # TAM Turquoise (highlight)
DARK_BG = "070B14"           # Ultra-dark background
CARD_BG = "0C1220"           # Card background
HEADER_BG = "111A2E"         # Header/section background
TEXT_PRIMARY = "E6EDF3"
TEXT_SECONDARY = "8B949E"
TEXT_MUTED = "4A5568"
GREEN = "22C55E"             # TAM Logo Green (success/growth)
RED = "EF4444"
ORANGE = "F59E0B"
WHITE = "FFFFFF"
BORDER_COLOR = "1E293B"


def _parse_metrics_from_section(section_text: str) -> list:
    """Extract key-value metric pairs from analysis text."""
    metrics = []
    patterns = [
        r"(?:Price|Market Cap|Revenue|Net Income|EPS|P/E|P/B|ROE|ROA|"
        r"Dividend Yield|Debt.to.Equity|Current Ratio|Beta|"
        r"Free Cash Flow|Operating Margin|Profit Margin|"
        r"Book Value|Total Assets|Total Debt|EBITDA|"
        r"52.Week High|52.Week Low|Volume|Shares Outstanding"
        r")[\s:]+\$?([A-Z]{3}\s)?[\d,]+\.?\d*[BMK%]?",
    ]
    # Simple key: value extraction
    for line in section_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        # Look for "Key: Value" or "Key | Value" patterns
        kv_match = re.match(
            r"^[\|\*\-\s]*([A-Za-z\s/\(\)&]+?)[\s]*[:\|]+[\s]*(.+?)[\s\|]*$",
            line
        )
        if kv_match:
            key = kv_match.group(1).strip().strip("*").strip()
            value = kv_match.group(2).strip().strip("*").strip("|").strip()
            if key and value and len(key) < 50 and len(value) < 80:
                metrics.append((key, value))
    return metrics


def generate_xlsx_report(
    company_name: str,
    ticker: str,
    sections: dict,
    charts: dict = None,
    output_dir: str = "output",
    sources=None,
) -> str:
    """Generate a branded XLSX workbook with analysis data.

    Returns the path to the generated file.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for XLSX generation. pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- Styles ----
    header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    accent_fill = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
    row_fill_even = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    row_fill_odd = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color=ACCENT_BLUE)
    subtitle_font = Font(name="Calibri", bold=True, size=11, color=TEXT_SECONDARY)
    body_font = Font(name="Calibri", size=10, color=TEXT_PRIMARY)
    muted_font = Font(name="Calibri", size=9, color=TEXT_MUTED)
    thin_border = Border(
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    date_str = datetime.now().strftime("%B %d, %Y")

    # ==============================
    # Sheet 1: Executive Summary
    # ==============================
    ws_summary = wb.create_sheet("Executive Summary")
    ws_summary.sheet_properties.tabColor = ACCENT_BLUE

    # Title row
    ws_summary.merge_cells("A1:F1")
    cell = ws_summary["A1"]
    cell.value = f"{company_name} ({ticker}) — Investment Research Report"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center")

    ws_summary.merge_cells("A2:F2")
    ws_summary["A2"].value = f"TAM Capital | {date_str} | Confidential"
    ws_summary["A2"].font = muted_font

    ws_summary.merge_cells("A3:F3")  # spacer

    # Executive summary text
    exec_text = sections.get("executive_summary", "No executive summary generated.")
    row = 4
    for para in exec_text.split("\n"):
        para = para.strip()
        if not para:
            continue
        ws_summary.merge_cells(f"A{row}:F{row}")
        c = ws_summary[f"A{row}"]
        c.value = para
        c.font = body_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws_summary.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws_summary.column_dimensions[col_letter].width = 18

    # ==============================
    # Sheet 2: Key Metrics Dashboard
    # ==============================
    ws_metrics = wb.create_sheet("Key Metrics")
    ws_metrics.sheet_properties.tabColor = TURQUOISE

    ws_metrics.merge_cells("A1:C1")
    ws_metrics["A1"].value = f"Key Financial Metrics — {company_name}"
    ws_metrics["A1"].font = title_font

    # Headers
    for col_idx, header_text in enumerate(["Metric", "Value", "Section"], start=1):
        c = ws_metrics.cell(row=3, column=col_idx, value=header_text)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # Extract metrics from all sections
    metric_row = 4
    for section_key in SECTION_ORDER:
        content = sections.get(section_key, "")
        if not content:
            continue
        section_name = SECTION_TITLES.get(section_key, section_key)
        extracted = _parse_metrics_from_section(content)
        for key, value in extracted[:30]:  # cap per section
            fill = row_fill_even if (metric_row % 2 == 0) else row_fill_odd
            c1 = ws_metrics.cell(row=metric_row, column=1, value=key)
            c1.font = body_font
            c1.fill = fill
            c1.border = thin_border

            c2 = ws_metrics.cell(row=metric_row, column=2, value=value)
            c2.font = body_font
            c2.fill = fill
            c2.alignment = Alignment(horizontal="right")
            c2.border = thin_border

            c3 = ws_metrics.cell(row=metric_row, column=3, value=section_name)
            c3.font = muted_font
            c3.fill = fill
            c3.border = thin_border

            metric_row += 1

    ws_metrics.column_dimensions["A"].width = 30
    ws_metrics.column_dimensions["B"].width = 22
    ws_metrics.column_dimensions["C"].width = 30

    # ==============================
    # Per-section analysis sheets
    # ==============================
    for section_key in SECTION_ORDER:
        content = sections.get(section_key, "")
        if not content:
            continue
        title = SECTION_TITLES.get(section_key, section_key.replace("_", " ").title())
        safe_title = title[:31]  # Excel sheet name limit

        ws = wb.create_sheet(safe_title)
        ws.sheet_properties.tabColor = DARK_BLUE

        ws.merge_cells("A1:E1")
        ws["A1"].value = title
        ws["A1"].font = title_font

        ws.merge_cells("A2:E2")
        ws["A2"].value = f"{company_name} | {date_str}"
        ws["A2"].font = muted_font

        row = 4
        for para in content.split("\n"):
            para = para.strip()
            if not para:
                row += 1
                continue
            ws.merge_cells(f"A{row}:E{row}")
            c = ws[f"A{row}"]
            # Detect headers
            if para.startswith("#") or para.startswith("**") or para.isupper():
                c.value = para.lstrip("#").strip("*").strip()
                c.font = subtitle_font
            else:
                c.value = para
                c.font = body_font
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        for col_letter in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col_letter].width = 22

    # ==============================
    # Sources sheet
    # ==============================
    if sources and len(sources) > 0:
        ws_src = wb.create_sheet("Sources")
        ws_src.sheet_properties.tabColor = TEXT_MUTED

        ws_src["A1"].value = "Sources & References"
        ws_src["A1"].font = title_font

        for col_idx, header_text in enumerate(["#", "Source", "URL", "Type"], start=1):
            c = ws_src.cell(row=3, column=col_idx, value=header_text)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border

        row = 4
        try:
            for i, src in enumerate(sources.sources, 1):
                ws_src.cell(row=row, column=1, value=i).font = muted_font
                ws_src.cell(row=row, column=2, value=src.get("title", "")).font = body_font
                ws_src.cell(row=row, column=3, value=src.get("url", "")).font = muted_font
                ws_src.cell(row=row, column=4, value=src.get("type", "")).font = muted_font
                row += 1
        except Exception:
            pass

        ws_src.column_dimensions["A"].width = 6
        ws_src.column_dimensions["B"].width = 40
        ws_src.column_dimensions["C"].width = 60
        ws_src.column_dimensions["D"].width = 15

    # ---- Save ----
    os.makedirs(output_dir, exist_ok=True)
    filename = report_filename(company_name, "xlsx")
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
